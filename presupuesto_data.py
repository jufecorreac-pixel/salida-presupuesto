# -*- coding: utf-8 -*-
"""Lectura del Excel Presupuesto.xlsx y utilidades de la app.

El Excel es la fuente de verdad: los precios y los items salen de ahi.
Los totales NO se leen del Excel, se recalculan sumando los items
(la formula de CENA - SANDWICHES omite el pan tajado).
"""

import json
import os

import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RUTA_EXCEL = os.path.join(BASE_DIR, "Presupuesto.xlsx")
RUTA_HOSPEDAJES = os.path.join(BASE_DIR, "hospedajes.json")

HOSPEDAJE_MAX = 350_000
HOSPEDAJE_PASO = 1_000


def cop(n):
    """Formato de pesos colombianos: 251336 -> '$251.336'."""
    return "${:,.0f}".format(round(n)).replace(",", ".")


def _nombre_visible(seccion):
    """'ALMUERZO - LOMO + PAPAS' -> 'Lomo + papas'."""
    resto = seccion.split("-", 1)[1].strip() if "-" in seccion else seccion.strip()
    return resto.capitalize()


def cargar_presupuesto(ruta=RUTA_EXCEL):
    """Devuelve el presupuesto agrupado por seccion.

    {
      "desayunos":     {"nombre", "items", "total"},
      "almuerzos":     [{"nombre", "items", "total"}, ...],
      "cenas":         [{"nombre", "items", "total"}, ...],
      "lavada":        {"nombre", "items", "total"},
      "hospedaje_ref": 251336,
    }
    """
    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb["Presupuesto"]

    # La columna A solo trae el nombre de seccion en algunas filas: se arrastra.
    secciones = {}
    orden = []
    actual = None
    for fila in range(2, ws.max_row + 1):
        concepto = ws.cell(fila, 1).value
        producto = ws.cell(fila, 2).value
        cantidad = ws.cell(fila, 3).value
        precio = ws.cell(fila, 4).value

        if concepto:
            actual = str(concepto).strip()
        if not producto or actual is None:
            continue
        producto = str(producto).strip()
        if producto.upper() == "TOTAL":
            continue  # los totales se recalculan aqui, no se leen del Excel
        if not isinstance(precio, (int, float)):
            continue

        if actual not in secciones:
            secciones[actual] = []
            orden.append(actual)
        secciones[actual].append(
            {
                "producto": producto,
                "cantidad": str(cantidad).strip() if cantidad else "",
                "precio": int(precio),
            }
        )
    wb.close()

    datos = {
        "desayunos": None,
        "almuerzos": [],
        "cenas": [],
        "lavada": None,
        "hospedaje_ref": 0,
    }

    for seccion in orden:
        items = secciones[seccion]
        bloque = {
            "nombre": _nombre_visible(seccion),
            "seccion": seccion,
            "items": items,
            "total": sum(i["precio"] for i in items),
        }
        clave = seccion.upper()
        if clave.startswith("DESAYUNO"):
            bloque["nombre"] = seccion.capitalize()
            datos["desayunos"] = bloque
        elif clave.startswith("ALMUERZO"):
            datos["almuerzos"].append(bloque)
        elif clave.startswith("CENA"):
            datos["cenas"].append(bloque)
        elif clave.startswith("CARRO"):
            bloque["nombre"] = items[0]["producto"] if items else "Lavada de carro"
            datos["lavada"] = bloque
        elif clave.startswith("HOSPEDAJE"):
            # Solo referencia: el valor que cuenta lo pone la barra de la app.
            datos["hospedaje_ref"] = bloque["total"]

    return datos


def cargar_hospedajes(ruta=RUTA_HOSPEDAJES):
    """Lista de Airbnb: [{"nombre": ..., "url": ...}, ...]."""
    if not os.path.exists(ruta):
        return []
    try:
        with open(ruta, "r", encoding="utf-8") as f:
            datos = json.load(f)
    except (json.JSONDecodeError, OSError):
        return []
    if not isinstance(datos, list):
        return []
    limpios = []
    for item in datos:
        if isinstance(item, dict):
            limpios.append(
                {
                    "nombre": str(item.get("nombre", "") or "").strip(),
                    "url": str(item.get("url", "") or "").strip(),
                }
            )
    return limpios


def guardar_hospedajes(lista, ruta=RUTA_HOSPEDAJES):
    """Guarda la lista de Airbnb, descartando filas totalmente vacias."""
    limpios = []
    for item in lista:
        nombre = str(item.get("nombre", "") or "").strip()
        url = str(item.get("url", "") or "").strip()
        if nombre or url:
            limpios.append({"nombre": nombre, "url": url})
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(limpios, f, ensure_ascii=False, indent=2)
    return limpios


if __name__ == "__main__":
    d = cargar_presupuesto()
    print("Hospedaje (referencia del Excel):", cop(d["hospedaje_ref"]))
    print("\nDesayunos (fijo):", cop(d["desayunos"]["total"]))
    for i in d["desayunos"]["items"]:
        print("   -", i["producto"], i["cantidad"], cop(i["precio"]))
    print("\nAlmuerzos:")
    for a in d["almuerzos"]:
        print("   -", a["nombre"], "->", cop(a["total"]), "(%d items)" % len(a["items"]))
    print("\nCenas:")
    for c in d["cenas"]:
        print("   -", c["nombre"], "->", cop(c["total"]), "(%d items)" % len(c["items"]))
    print("\nLavada:", d["lavada"]["nombre"], "->", cop(d["lavada"]["total"]))
